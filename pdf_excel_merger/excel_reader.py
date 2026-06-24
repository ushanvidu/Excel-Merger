"""Read a target Excel template: list sheets and locate header columns.

The template is the file the user wants *filled*. We need its column names and
their exact positions (row + column index) so the merger can write values into
the right cells while leaving the rest of the workbook untouched.
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .utils import clean_cell


@dataclass
class TargetColumn:
    """One column in the template that can receive data."""

    name: str          # header text shown to the user
    index: int         # 1-based column index (openpyxl convention)
    letter: str        # e.g. "A", "B"


@dataclass
class TemplateInfo:
    """Everything the UI/merger needs about one sheet of the template."""

    sheet_name: str
    header_row: int
    first_data_row: int
    columns: list[TargetColumn]
    last_filled_row: int  # last row with any data (for append mode)

    @property
    def column_names(self) -> list[str]:
        return [c.name for c in self.columns]


def list_sheets(path: str) -> list[str]:
    wb = load_workbook(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def detect_header_row(path: str, sheet_name: str, max_scan: int = 20) -> int:
    """Best-effort guess of the header row: the first row with >=2 non-empty,
    mostly-text cells. Falls back to row 1."""
    wb = load_workbook(path, read_only=True)
    try:
        ws = wb[sheet_name]
        for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan), start=1):
            values = [clean_cell(c.value) for c in row]
            non_empty = [v for v in values if v]
            if len(non_empty) >= 2:
                text_like = sum(1 for v in non_empty if not v.replace(".", "").isdigit())
                if text_like >= max(1, len(non_empty) // 2):
                    return r
        return 1
    finally:
        wb.close()


def read_template(path: str, sheet_name: str, header_row: int) -> TemplateInfo:
    """Read header columns and data extent for one sheet."""
    wb = load_workbook(path)
    try:
        ws: Worksheet = wb[sheet_name]
        columns: list[TargetColumn] = []
        for cell in ws[header_row]:
            name = clean_cell(cell.value)
            if name == "":
                continue
            columns.append(
                TargetColumn(
                    name=name,
                    index=cell.column,
                    letter=get_column_letter(cell.column),
                )
            )
        last_filled = _last_filled_row(ws, header_row, columns)
        return TemplateInfo(
            sheet_name=sheet_name,
            header_row=header_row,
            first_data_row=header_row + 1,
            columns=columns,
            last_filled_row=last_filled,
        )
    finally:
        wb.close()


def column_by_name(info: TemplateInfo, name: str) -> TargetColumn | None:
    for col in info.columns:
        if col.name == name:
            return col
    return None


def read_key_values(
    source, info: TemplateInfo, key_column_name: str
) -> list[tuple[int, str]]:
    """Read ``(row_number, key_text)`` pairs from the master sheet's key column.

    ``source`` is a path or a bytes/BytesIO of the workbook. Blank keys are
    skipped. Used to build a ``site_matching.SiteIndex``.
    """
    col = column_by_name(info, key_column_name)
    if col is None:
        return []
    wb = load_workbook(source, read_only=True)
    try:
        ws = wb[info.sheet_name]
        out: list[tuple[int, str]] = []
        for r in range(info.first_data_row, ws.max_row + 1):
            value = clean_cell(ws.cell(row=r, column=col.index).value)
            if value != "":
                out.append((r, value))
        return out
    finally:
        wb.close()


def _last_filled_row(ws: Worksheet, header_row: int, columns: list[TargetColumn]) -> int:
    """Last row (>= header_row) that has any value in the mapped columns."""
    if not columns:
        return header_row
    col_indexes = [c.index for c in columns]
    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        for ci in col_indexes:
            if clean_cell(ws.cell(row=r, column=ci).value) != "":
                last = r
                break
    return last
