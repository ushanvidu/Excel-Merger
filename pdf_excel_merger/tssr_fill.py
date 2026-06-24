"""Fill the master 'RF Parameter' sheet's TSSR columns from TSSR PDF data.

Decisions confirmed with the user:
  - Source     : the **Solution** tables (Solution Antenna/Sector Details).
  - Match key  : **Sector + Band** (band->PDF-token mapping is configurable).
  - TSSR Status: left blank.

Join, per master row (one Sector x Tech x Band):
  1. Determine the sector number from the master 'Sector' (S1->1); if that cell
     is broken (#REF!/blank), recover it from 'Antenna azimuth_Mbitel' which
     matches the PDF 'Antenna Direction'.
  2. Match a Solution **Sector Details** row by sector + band token  -> tilts
     (Electrical Tilt 01, Mechanical Tilt) and the Antenna No.
  3. Look up the Solution **Antenna Details** row by sector (+ that Antenna No,
     preferring one whose ports mention the band)  -> azimuth, height, type.
Only mapped cells are written; everything else in the workbook is untouched.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

from .tssr_extractor import ANTENNA_COLS, SECTOR_COLS, TssrResult, merge_results
from .utils import clean_cell, coerce_value

# Master columns the default TSSR mapping fills (kept for the accuracy harness).
MASTER_TSSR_COLUMNS = [
    "Antenna_TSSR", "Antenna height_TSSR", "Antenna azimuth_TSSR",
    "Mechanical tilt_TSSR", "Electrical tilt_TSSR", "TSSR Status",
]

# Source fields available to map from, per Solution table (for the UI).
ANTENNA_FIELDS = [c for c in ANTENNA_COLS if c not in ("Sector No", "Antenna No")]
SECTOR_FIELDS = [c for c in SECTOR_COLS if c not in ("Sector No", "Antenna No")]

# Default mapping: master column -> (which Solution table, which field, coercion).
# Fully overridable from the UI so the app fits any master + any required columns.
DEFAULT_FIELD_MAP = [
    {"target": "Antenna_TSSR", "section": "antenna", "field": "Antenna Type", "kind": "text"},
    {"target": "Antenna height_TSSR", "section": "antenna", "field": "Antenna Height (m)", "kind": "number"},
    {"target": "Antenna azimuth_TSSR", "section": "antenna", "field": "Antenna Direction", "kind": "number"},
    {"target": "Mechanical tilt_TSSR", "section": "sector", "field": "Mechanical Tilt", "kind": "number"},
    {"target": "Electrical tilt_TSSR", "section": "sector", "field": "Electrical Tilt 01", "kind": "number"},
]

# Default master-Band -> regex matched against the Solution Sector 'Sector' token
# (e.g. "L850_1", "1800|L1800_P", "UMTS_W", "900_A", "L2.3_sec1", "L2.6_sec1").
DEFAULT_BAND_MAP = {
    "700": r"700|n28",
    "850": r"l?\s*850",
    "900": r"900",
    "1800": r"1800|l1800",
    "2100": r"umts|2100|l2100|w2100",
    "2300": r"l2\.?3|2300",
    "2600": r"l2\.?6|2600",
    "3500": r"5g|n78|3500|n3500",
}

_SECTOR_NUM_RE = re.compile(r"(?i)s(?:ector)?\s*0*(\d+)")
_ANTENNA_NUM_RE = re.compile(r"(?i)a(?:ntenna)?\s*0*(\d+)")


@dataclass
class FillReport:
    rows_filled: int = 0
    rows_unmatched: int = 0
    cells_written: dict[str, int] = field(default_factory=dict)
    details: list[dict] = field(default_factory=list)          # filled rows
    unmatched_details: list[dict] = field(default_factory=list)  # rows not filled
    conflicts: list[dict] = field(default_factory=list)        # sector signal disagreements
    sites_seen: set[str] = field(default_factory=set)
    sites_no_pdf: set[str] = field(default_factory=set)


def sector_num(value) -> int | None:
    m = _SECTOR_NUM_RE.search(clean_cell(value))
    return int(m.group(1)) if m else None


def antenna_num(value) -> int | None:
    """Parse the antenna number from an 'Antenna N' value (sector_num won't)."""
    m = _ANTENNA_NUM_RE.search(clean_cell(value))
    return int(m.group(1)) if m else None


# Letter suffixes that telecom cell tokens use to encode the sector ordinal.
_SUFFIX_ORD = {"a": 1, "b": 2, "c": 3, "d": 4,
               "p": 1, "q": 2, "r": 3, "s": 4,
               "w": 1, "x": 2, "y": 3, "z": 4}
_SUFFIX_NUM_RE = re.compile(r"(?:_|sec)0*([0-9]+)$", re.I)
_SUFFIX_LETTER_RE = re.compile(r"_([A-Za-z])$")


def suffix_ordinal(token) -> int | None:
    """1-based sector ordinal independently encoded in a Solution 'Sector' token.

    Examples: 'L850_1'->1, '1800|L1800_P'->1, 'UMTS_W'->1, 'L2.6_sec3'->3,
    '900_C'->3. This is a signal independent of the row's 'Sector No', used to
    cross-check the sector a row was matched to. Returns None when there is no
    recognizable suffix (no signal, so no false conflict).
    """
    t = clean_cell(token)
    m = _SUFFIX_NUM_RE.search(t)
    if m:
        n = int(m.group(1))
        return n if 1 <= n <= 8 else None
    m = _SUFFIX_LETTER_RE.search(t)
    if m:
        return _SUFFIX_ORD.get(m.group(1).lower())
    return None


def _band_token(band) -> str:
    """Normalize a band value, so '1800.0' (float cell) matches the '1800' key."""
    s = clean_cell(band)
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return s


def _direction_to_sector(result: TssrResult) -> dict[str, int]:
    """Map a PDF 'Antenna Direction' value -> sector number (for #REF! recovery).

    Built from both Existing and Solution antenna tables, since the master's
    'Antenna azimuth_Mbitel' tracks the existing on-site direction.
    """
    out: dict[str, int] = {}
    for section in ("Antenna Details", "Solution Antenna Details"):
        df = result.sections.get(section)
        if df is None:
            continue
        for _, r in df.iterrows():
            sn = sector_num(r["Sector No"])
            d = clean_cell(r["Antenna Direction"])
            if sn and d and d not in out:
                out[d] = sn
    return out


def _pdf_sector_set(result: TssrResult) -> set[int]:
    """All sector numbers present in a site's PDF (to detect 0- vs 1-based)."""
    out: set[int] = set()
    for section in result.sections.values():
        if "Sector No" in section.columns:
            for v in section["Sector No"]:
                n = sector_num(v)
                if n is not None:
                    out.add(n)
    return out


def _match_sector_row(df, sn: int, band_rx: re.Pattern, target_ordinal: int | None = None):
    """Solution Sector Details row for this sector whose token matches the band.

    When several rows in the sector match (e.g. duplicate band entries), prefer
    the one whose token suffix agrees with the sector (``target_ordinal``, the
    1-based sector label); otherwise take the first match.
    """
    matches = [r for _, r in df.iterrows()
               if sector_num(r["Sector No"]) == sn and band_rx.search(clean_cell(r["Sector"]))]
    if not matches:
        return None
    if target_ordinal is not None and len(matches) > 1:
        for r in matches:
            if suffix_ordinal(r["Sector"]) == target_ordinal:
                return r
    return matches[0]


def _match_antenna_row(df, sn: int, antenna_no: str | None, band: str):
    """Solution Antenna Details row for this sector, preferring antenna_no and a
    port that mentions the band; falls back to the first sector match."""
    candidates = [r for _, r in df.iterrows() if sector_num(r["Sector No"]) == sn]
    if not candidates:
        return None
    pool = candidates
    an = antenna_num(antenna_no) if antenna_no else None
    if an is not None:
        named = [r for r in pool if antenna_num(r["Antenna No"]) == an]
        if named:
            pool = named
    if band:
        ports = [r for r in pool
                 if any(band in clean_cell(r[f"Port 0{i}"]) for i in range(1, 6))]
        if ports:
            return ports[0]
    return pool[0]


def _cell(ws_values, ws_raw, row: int, c: int | None):
    """Read a cell, preferring the cached value over a formula string.

    ``ws_values`` is from a ``data_only=True`` load (cached results), ``ws_raw``
    from a normal load. This keeps key/azimuth reads working when the master uses
    formulas (e.g. a broken =XLOOKUP in the Sector column).
    """
    if c is None:
        return None
    v = ws_values.cell(row, c).value
    if v not in (None, ""):
        return v
    raw = ws_raw.cell(row, c).value
    if isinstance(raw, str) and raw.startswith("="):
        return None  # formula with no cached value -> unusable as a key
    return raw


def fill_master(
    master_bytes: bytes,
    results: list[TssrResult],
    *,
    sheet_name: str = "RF Parameter",
    header_row: int = 1,
    site_col: str = "Site Name",
    sector_col: str = "Sector",
    azimuth_col: str = "Antenna azimuth_Mbitel",
    band_col: str = "Band",
    source: str = "Solution",
    band_map: dict[str, str] | None = None,
    field_map: list[dict] | None = None,
) -> tuple[bytes, FillReport]:
    """Fill the master sheet and return (filled_xlsx_bytes, report).

    ``field_map`` maps each master column to a Solution field — a list of
    ``{"target", "section": "antenna"|"sector", "field", "kind"}``. Defaults to
    the TSSR mapping but is fully overridable, so the app fits any master and any
    set of columns to fill. Site rows are matched by Site + Sector + Band.
    """
    band_map = band_map or DEFAULT_BAND_MAP
    field_map = field_map or DEFAULT_FIELD_MAP
    band_rx = {str(b): re.compile(rx, re.I) for b, rx in band_map.items()}
    by_site = {r.site_id.upper(): r for r in merge_results(results)}
    report = FillReport()

    prefix = "Solution " if str(source).lower().startswith("sol") else ""
    sec_section, ant_section = prefix + "Sector Details", prefix + "Antenna Details"

    wb = load_workbook(io.BytesIO(master_bytes))                 # for writing (keeps formulas)
    wb_values = load_workbook(io.BytesIO(master_bytes), data_only=True)  # for reading
    ws, ws_values = wb[sheet_name], wb_values[sheet_name]

    header = {clean_cell(c.value): c.column for c in ws[header_row]}
    targets = [spec["target"] for spec in field_map]
    required = [site_col, band_col, *targets]
    missing = [c for c in required if c not in header]
    if missing:
        raise ValueError(f"Master sheet is missing required column(s): {missing}")
    col = header  # name -> column index
    col_site, col_band = col[site_col], col[band_col]
    col_sector, col_az = col.get(sector_col), col.get(azimuth_col)

    def read(ws_row, name_col):
        return _cell(ws_values, ws, ws_row, name_col)

    # Pre-pass: per-site azimuth -> sector map (handles broken Sector cells and
    # sites whose surveyed direction differs from the master azimuth).
    site_rows: dict[str, list[tuple[int | None, str]]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        s = clean_cell(read(r, col_site))
        if not s or s.upper() not in by_site:
            continue
        site_rows.setdefault(s, []).append((
            sector_num(read(r, col_sector)),
            clean_cell(read(r, col_az)),
        ))

    az_maps: dict[str, dict[str, int]] = {}
    offsets: dict[str, int] = {}
    for s, rows in site_rows.items():
        result = by_site[s.upper()]
        dmap = _direction_to_sector(result)
        sset = _pdf_sector_set(result)
        offset = -1 if (sset and min(sset) == 0) else 0
        for raw, az in rows:
            if az and az not in dmap and raw is not None:
                sec = raw + offset
                if not sset or sec in sset:
                    dmap[az] = sec
        az_maps[s], offsets[s] = dmap, offset

    fills = {t: 0 for t in targets}

    for row in range(header_row + 1, ws.max_row + 1):
        site = clean_cell(read(row, col_site))
        if not site:
            continue
        result = by_site.get(site.upper())
        report.sites_seen.add(site)
        if result is None:
            report.sites_no_pdf.add(site)
            continue

        # Resolve the PDF sector: azimuth map first, then master Sector + offset.
        azimuth = clean_cell(read(row, col_az))
        sn = az_maps.get(site, {}).get(azimuth)
        if sn is None:
            raw = sector_num(read(row, col_sector))
            if raw is not None:
                sn = raw + offsets.get(site, 0)
        band = _band_token(read(row, col_band))
        if sn is None or not band:
            report.rows_unmatched += 1
            report.unmatched_details.append({
                "row": row, "site": site, "sector": sn or "", "band": band,
                "reason": "sector not resolved" if sn is None else "no band value",
            })
            continue

        sec_df = result.sections.get(sec_section)
        ant_df = result.sections.get(ant_section)
        rx = band_rx.get(band)
        target_ord = sn - offsets.get(site, 0)  # 1-based sector label
        sec_row = (_match_sector_row(sec_df, sn, rx, target_ord)
                   if rx is not None and sec_df is not None else None)
        antenna_no = clean_cell(sec_row["Antenna No"]) if sec_row is not None else None
        ant_row = _match_antenna_row(ant_df, sn, antenna_no, band) if ant_df is not None else None

        # Cross-signal check: token suffix vs resolved sector.
        conflict = ""
        if sec_row is not None:
            token = clean_cell(sec_row["Sector"])
            so = suffix_ordinal(token)
            if so is not None and so + offsets.get(site, 0) != sn:
                conflict = f"token '{token}' implies sector {so}, resolved {target_ord}"
                report.conflicts.append({
                    "row": row, "site": site, "band": band, "resolved_sector": target_ord,
                    "token": token, "token_sector": so, "detail": conflict,
                })

        section_rows = {"antenna": ant_row, "sector": sec_row}
        values: dict[str, object] = {}
        for spec in field_map:
            src_row = section_rows.get(spec["section"])
            if src_row is None or spec["field"] not in src_row.index:
                continue
            values[spec["target"]] = coerce_value(src_row[spec["field"]], spec.get("kind", "auto"))

        wrote = False
        for name, val in values.items():
            if val is None or (isinstance(val, str) and val == ""):
                continue
            ws.cell(row, col[name], val)
            fills[name] += 1
            wrote = True

        if not wrote:
            report.rows_unmatched += 1
            report.unmatched_details.append({
                "row": row, "site": site, "sector": sn, "band": band,
                "reason": "no Solution row for this sector+band",
            })
            continue

        report.rows_filled += 1
        report.details.append({
            "row": row, "site": site, "sector": target_ord, "band": band,
            "token": clean_cell(sec_row["Sector"]) if sec_row is not None else "",
            "antenna": antenna_no or "",
            **{t: clean_cell(values.get(t, "")) for t in targets},
            "status": "review" if conflict else "filled",
            "conflict": conflict,
        })

    report.cells_written = {k: v for k, v in fills.items() if v}
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), report
