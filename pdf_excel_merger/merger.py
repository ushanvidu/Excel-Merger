"""Write mapped source data into the Excel template.

The template workbook is loaded with openpyxl and only the target cells are
written, so existing formatting, formulas and other sheets are preserved. New
data rows copy the cell style of the template's first data row when present.
"""

from __future__ import annotations

import copy
import io
from dataclasses import dataclass, field

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from . import excel_reader
from .excel_reader import TemplateInfo
from .pdf_extractor import SOURCE_COLUMN
from .site_matching import DEFAULT_FUZZY_THRESHOLD, SiteIndex
from .utils import clean_cell, coerce_value, infer_column_kind


@dataclass
class MergeReport:
    rows_written: int = 0
    start_row: int = 0
    end_row: int = 0
    mode: str = "append"
    blanks_per_column: dict[str, int] = field(default_factory=dict)
    unmapped_targets: list[str] = field(default_factory=list)
    coercion_notes: list[str] = field(default_factory=list)


def _copy_style(src: Cell, dst: Cell) -> None:
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)


def merge(
    template_bytes: bytes,
    template_info: TemplateInfo,
    source_df: pd.DataFrame,
    mapping: dict[str, str | None],
    *,
    mode: str = "append",
    column_kinds: dict[str, str] | None = None,
) -> tuple[bytes, MergeReport]:
    """Fill the template and return (xlsx_bytes, report).

    Parameters
    ----------
    template_bytes : the uploaded .xlsx as bytes.
    template_info  : result of excel_reader.read_template for the chosen sheet.
    source_df      : combined extracted data (columns = source field names).
    mapping        : {target_column_name: source_column_name or None}.
    mode           : "append" writes below existing data; "overwrite" starts at
                     the first data row.
    column_kinds   : optional {target_column_name: "auto|text|number|date"};
                     missing entries are inferred from the source data.
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb[template_info.sheet_name]
    report = MergeReport(mode=mode)

    col_by_name = {c.name: c for c in template_info.columns}
    active_mapping = {
        tgt: src for tgt, src in mapping.items()
        if src and src in source_df.columns and tgt in col_by_name
    }
    report.unmapped_targets = [
        c.name for c in template_info.columns if c.name not in active_mapping
    ]

    # Resolve a coercion kind for each mapped target column.
    kinds: dict[str, str] = {}
    for tgt, src in active_mapping.items():
        kind = (column_kinds or {}).get(tgt, "auto")
        if kind == "auto":
            kind = infer_column_kind(source_df[src])
        kinds[tgt] = kind

    if mode == "overwrite":
        start_row = template_info.first_data_row
    else:
        start_row = max(template_info.last_filled_row + 1, template_info.first_data_row)

    style_row = template_info.first_data_row  # template's first data row = style source
    blanks = {tgt: 0 for tgt in active_mapping}

    n = len(source_df)
    for i in range(n):
        excel_row = start_row + i
        src_record = source_df.iloc[i]
        for tgt, src in active_mapping.items():
            col = col_by_name[tgt]
            raw = src_record[src]
            value = coerce_value(raw, kinds[tgt])
            if value is None or (isinstance(value, str) and value == ""):
                blanks[tgt] += 1
            cell = ws.cell(row=excel_row, column=col.index, value=value)
            # Inherit template formatting for freshly written rows.
            if excel_row > style_row:
                _copy_style(ws.cell(row=style_row, column=col.index), cell)

    report.rows_written = n
    report.start_row = start_row
    report.end_row = start_row + n - 1 if n else start_row
    report.blanks_per_column = {k: v for k, v in blanks.items() if v}

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), report


# ---------------------------------------------------------------------------
# Key-based update: match each source record to a master-sheet row by site key
# and fill that row's mapped columns (the network-sites workflow).
# ---------------------------------------------------------------------------

@dataclass
class KeyMergeReport:
    rows_updated: int = 0
    rows_appended: int = 0
    matched: list[dict] = field(default_factory=list)    # per source record
    unmatched: list[dict] = field(default_factory=list)
    fields_filled: dict[str, int] = field(default_factory=dict)
    fuzzy_matches: int = 0


def preview_key_matches(
    template_bytes: bytes,
    template_info: TemplateInfo,
    source_df: pd.DataFrame,
    *,
    main_key_col: str,
    source_key_col: str,
    allow_fuzzy: bool = True,
    fuzzy_threshold: float = DEFAULT_FUZZY_THRESHOLD,
) -> pd.DataFrame:
    """Show how each source record would match a master row (no file changes)."""
    entries = excel_reader.read_key_values(
        io.BytesIO(template_bytes), template_info, main_key_col
    )
    index = SiteIndex(entries)
    rows = []
    for _, record in source_df.iterrows():
        raw_key = clean_cell(record.get(source_key_col, ""))
        res = index.match(raw_key, allow_fuzzy=allow_fuzzy, threshold=fuzzy_threshold)
        rows.append({
            "Document": record.get(SOURCE_COLUMN, ""),
            "Site key (source)": raw_key,
            "Matched site (master)": res.matched_value,
            "Master row": res.row if res.row else "",
            "Match": res.method,
            "Score": f"{res.score:.0f}" if res.row else "",
        })
    return pd.DataFrame(rows)


def merge_by_key(
    template_bytes: bytes,
    template_info: TemplateInfo,
    source_df: pd.DataFrame,
    field_mapping: dict[str, str | None],
    *,
    main_key_col: str,
    source_key_col: str,
    on_unmatched: str = "skip",          # "skip" | "append"
    allow_fuzzy: bool = True,
    fuzzy_threshold: float = DEFAULT_FUZZY_THRESHOLD,
    column_kinds: dict[str, str] | None = None,
) -> tuple[bytes, KeyMergeReport]:
    """Update master-sheet rows by matching each source record's site key.

    For every source record the site key is matched to a master row; the mapped
    fields are then written into that row. Unmatched records are skipped or
    appended as new sites depending on ``on_unmatched``.
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb[template_info.sheet_name]
    report = KeyMergeReport()

    col_by_name = {c.name: c for c in template_info.columns}
    key_col = col_by_name.get(main_key_col)
    if key_col is None:
        raise ValueError(f"Key column '{main_key_col}' not found in template.")

    # Active field mapping excludes the key column and unusable entries.
    active = {
        tgt: src for tgt, src in field_mapping.items()
        if src and src in source_df.columns and tgt in col_by_name and tgt != main_key_col
    }

    kinds: dict[str, str] = {}
    for tgt, src in active.items():
        kind = (column_kinds or {}).get(tgt, "auto")
        if kind == "auto":
            kind = infer_column_kind(source_df[src])
        kinds[tgt] = kind

    entries = excel_reader.read_key_values(
        io.BytesIO(template_bytes), template_info, main_key_col
    )
    index = SiteIndex(entries)

    style_row = template_info.first_data_row
    next_append_row = max(template_info.last_filled_row + 1, template_info.first_data_row)
    filled = {tgt: 0 for tgt in active}

    for _, record in source_df.iterrows():
        doc = clean_cell(record.get(SOURCE_COLUMN, ""))
        raw_key = clean_cell(record.get(source_key_col, ""))
        res = index.match(raw_key, allow_fuzzy=allow_fuzzy, threshold=fuzzy_threshold)

        target_row = res.row
        appended = False
        if target_row is None:
            if on_unmatched == "append" and raw_key:
                target_row = next_append_row
                next_append_row += 1
                appended = True
                key_cell = ws.cell(row=target_row, column=key_col.index, value=raw_key)
                _copy_style(ws.cell(row=style_row, column=key_col.index), key_cell)
            else:
                report.unmatched.append({"document": doc, "site_key": raw_key})
                continue

        for tgt, src in active.items():
            col = col_by_name[tgt]
            value = coerce_value(record[src], kinds[tgt])
            cell = ws.cell(row=target_row, column=col.index, value=value)
            if appended and target_row > style_row:
                _copy_style(ws.cell(row=style_row, column=col.index), cell)
            if not (value is None or (isinstance(value, str) and value == "")):
                filled[tgt] += 1

        if appended:
            report.rows_appended += 1
        else:
            report.rows_updated += 1
            if res.method == "fuzzy":
                report.fuzzy_matches += 1
        report.matched.append({
            "document": doc,
            "site_key": raw_key,
            "matched_value": res.matched_value or raw_key,
            "row": target_row,
            "method": "appended" if appended else res.method,
            "score": res.score,
        })

    report.fields_filled = {k: v for k, v in filled.items() if v}

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), report


def preview_merge(
    template_info: TemplateInfo,
    source_df: pd.DataFrame,
    mapping: dict[str, str | None],
    *,
    rows: int = 20,
) -> pd.DataFrame:
    """Build a preview DataFrame shaped like the target columns (no file I/O)."""
    data = {}
    for col in template_info.columns:
        src = mapping.get(col.name)
        if src and src in source_df.columns:
            data[col.name] = source_df[src].head(rows).map(clean_cell).tolist()
        else:
            data[col.name] = [""] * min(rows, len(source_df))
    return pd.DataFrame(data)
