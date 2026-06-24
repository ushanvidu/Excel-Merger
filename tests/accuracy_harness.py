"""Golden-set accuracy harness for the TSSR -> RF Parameter fill.

Runs the real pipeline (extract the golden sites' PDFs + fill the master), then
compares the written TSSR cells against a hand-verified ground truth
(`tests/golden/tssr_golden.csv`). Reports per-field accuracy / precision / recall
and lists every mismatch, so accuracy is a number you can track on each change.

Run:  python -m tests.accuracy_harness
Exits non-zero if overall accuracy drops below ACCURACY_GATE (regression guard).
"""

from __future__ import annotations

import csv
import io
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from pdf_excel_merger.tssr_extractor import extract_tssr  # noqa: E402
from pdf_excel_merger.tssr_fill import MASTER_TSSR_COLUMNS, fill_master  # noqa: E402

SAMPLES = ROOT / "sample_data"
GOLDEN = ROOT / "tests" / "golden" / "tssr_golden.csv"
MASTER = SAMPLES / "Noding Config_V1.1.4.xlsx"

# Map golden Site -> its PDF in sample_data.
SITE_PDFS = {
    "GMALG1": "TSSRBTS5GZT of 5G Expansion Phase 1 - GMALG1 (1).pdf",
    "GMDAD1": "TSSRBTS5GZT_of_5G_Expansion_Phase_1_-_GMDAD1.pdf",
}

SCORED_FIELDS = MASTER_TSSR_COLUMNS[:-1]  # all TSSR cols except "TSSR Status"
ACCURACY_GATE = 0.90  # overall accuracy below this fails the run


def _norm(value) -> str:
    """Canonicalize a cell for comparison (numbers compare by value)."""
    if value is None:
        return ""
    s = str(value).strip()
    if s == "":
        return ""
    try:
        f = float(s)
        return str(int(f)) if f.is_integer() else str(f)
    except ValueError:
        return s


@dataclass
class FieldStat:
    checked: int = 0
    correct: int = 0       # expected==actual (incl. both blank)
    wrong: int = 0         # expected val, actual different non-blank val
    missing: int = 0       # expected val, actual blank
    spurious: int = 0      # expected blank, actual val
    mismatches: list = field(default_factory=list)
    _tp: int = 0           # true positives (expected non-blank, matched)

    def accuracy(self) -> float:
        return self.correct / self.checked if self.checked else 1.0

    def precision(self) -> float:
        denom = self._tp + self.wrong + self.spurious
        return self._tp / denom if denom else 1.0

    def recall(self) -> float:
        denom = self._tp + self.wrong + self.missing
        return self._tp / denom if denom else 1.0


def load_golden() -> list[dict]:
    rows = []
    with open(GOLDEN, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(r for r in f if not r.lstrip().startswith("#")):
            rows.append(row)
    return rows


def run() -> int:
    if not MASTER.exists():
        print(f"Master fixture missing: {MASTER} — skipping accuracy harness.")
        return 0

    golden = load_golden()
    sites = sorted({g["Site"] for g in golden})
    results = []
    for site in sites:
        pdf = SAMPLES / SITE_PDFS[site]
        if not pdf.exists():
            print(f"PDF fixture missing for {site}: {pdf} — skipping.")
            continue
        results.append(extract_tssr(pdf.read_bytes(), filename=pdf.name))

    out_bytes, _ = fill_master(MASTER.read_bytes(), results)
    ws = load_workbook(io.BytesIO(out_bytes))["RF Parameter"]
    hdr = {c.value: c.column for c in ws[1]}

    # Index master rows by (Site, Tech, Band, Azimuth) -> list of written field dicts.
    written = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        site = _norm(ws.cell(r, hdr["Site Name"]).value)
        if site not in SITE_PDFS:
            continue
        key = (site, _norm(ws.cell(r, hdr["Tech"]).value),
               _norm(ws.cell(r, hdr["Band"]).value),
               _norm(ws.cell(r, hdr["Antenna azimuth_Mbitel"]).value))
        written[key].append({c: ws.cell(r, hdr[c]).value for c in SCORED_FIELDS})

    stats = {c: FieldStat() for c in SCORED_FIELDS}
    unmatched_keys = []
    for g in golden:
        key = (g["Site"], _norm(g["Tech"]), _norm(g["Band"]), _norm(g["Azimuth"]))
        rows = written.get(key)
        if not rows:
            unmatched_keys.append(key)
            continue
        for actual in rows:  # duplicate master rows each get scored
            for col in SCORED_FIELDS:
                exp_raw = g[col].strip()
                if exp_raw == "?":
                    continue
                st = stats[col]
                st.checked += 1
                expected = "" if exp_raw == "BLANK" else _norm(exp_raw)
                got = _norm(actual[col])
                if expected == got:
                    st.correct += 1
                    if expected != "":
                        st._tp += 1
                    continue
                st.mismatches.append({"key": key, "expected": expected or "(blank)",
                                      "got": got or "(blank)"})
                if expected and got:
                    st.wrong += 1
                elif expected and not got:
                    st.missing += 1
                else:
                    st.spurious += 1

    # ---- Report ----
    print("=" * 74)
    print(f"TSSR FILL ACCURACY  (sites: {', '.join(sites)})")
    print("=" * 74)
    print(f"{'Field':<24}{'checked':>8}{'acc':>8}{'prec':>8}{'recall':>8}"
          f"{'wrong':>7}{'miss':>6}{'spur':>6}")
    tot_checked = tot_correct = 0
    for col in SCORED_FIELDS:
        s = stats[col]
        tot_checked += s.checked
        tot_correct += s.correct
        print(f"{col:<24}{s.checked:>8}{s.accuracy():>8.2%}{s.precision():>8.2%}"
              f"{s.recall():>8.2%}{s.wrong:>7}{s.missing:>6}{s.spurious:>6}")
    overall = tot_correct / tot_checked if tot_checked else 1.0
    print("-" * 74)
    print(f"{'OVERALL':<24}{tot_checked:>8}{overall:>8.2%}")

    mismatches = [(c, m) for c in SCORED_FIELDS for m in stats[c].mismatches]
    if mismatches:
        print(f"\nMISMATCHES ({len(mismatches)}):")
        for col, m in mismatches:
            site, tech, band, az = m["key"]
            print(f"  {site} {tech}/{band}@az{az:<4} {col:<22} "
                  f"expected={m['expected']!r} got={m['got']!r}")
    if unmatched_keys:
        print(f"\nGolden rows with no master match ({len(unmatched_keys)}): {unmatched_keys}")

    print()
    if overall < ACCURACY_GATE:
        print(f"FAIL: overall accuracy {overall:.2%} < gate {ACCURACY_GATE:.0%}")
        return 1
    print(f"PASS: overall accuracy {overall:.2%} >= gate {ACCURACY_GATE:.0%}")
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
