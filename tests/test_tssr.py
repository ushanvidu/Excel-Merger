"""End-to-end test for the TSSR workflow against the real sample files.

Skips gracefully if the real samples aren't present. Run:
    python -m tests.test_tssr
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from pdf_excel_merger.tssr_extractor import SECTION_SCHEMAS, extract_tssr, to_workbook_bytes  # noqa: E402
from pdf_excel_merger.tssr_fill import MASTER_TSSR_COLUMNS, fill_master  # noqa: E402

SAMPLES = ROOT / "sample_data"
PDF = SAMPLES / "TSSRBTS5GZT of 5G Expansion Phase 1 - GMALG1 (1).pdf"
MASTER = SAMPLES / "Sample " / "Noding Config_V1.1.4.xlsx"


def _fail(msg: str) -> None:
    raise AssertionError(msg)


def main() -> int:
    if not PDF.exists() or not MASTER.exists():
        print("Real TSSR samples not found — skipping (this test needs them).")
        return 0

    # 1. Extract the four topic sections.
    res = extract_tssr(PDF.read_bytes(), filename=PDF.name)
    counts = {k: len(res.sections[k]) for k in SECTION_SCHEMAS}
    print("Site:", res.site_id, "| section counts:", counts)
    if res.site_id != "GMALG1":
        _fail(f"site id parse wrong: {res.site_id}")
    if counts["Antenna Details"] < 5 or counts["Sector Details"] < 8:
        _fail(f"too few rows extracted: {counts}")
    # SB1800 antenna row must be classified as antenna, not sector
    if res.sections["Sector Details"]["Sector"].astype(str).str.contains("SB").any():
        _fail("SB antenna row leaked into Sector Details")
    # Geometry reconstruction must capture the full band token (no truncation):
    # the per-sector suffix _P/_Q/_R was previously dropped by extract_tables.
    sol_tokens = set(res.sections["Solution Sector Details"]["Sector"].astype(str))
    if not any(t.endswith(("_P", "_Q", "_R")) for t in sol_tokens):
        _fail(f"band-token suffix truncated (geometry regressed): {sorted(sol_tokens)}")

    # 2. Multi-sheet workbook export has the expected topic sheets.
    wb = load_workbook(io.BytesIO(to_workbook_bytes([res])))
    expected = {"Summary", *SECTION_SCHEMAS.keys(), "Extraction Log"}
    if not expected.issubset(set(wb.sheetnames)):
        _fail(f"workbook missing sheets: {expected - set(wb.sheetnames)}")
    print("Workbook sheets:", wb.sheetnames)

    # 3. Fill the master RF Parameter TSSR columns for GMALG1.
    out, report = fill_master(MASTER.read_bytes(), [res])
    print(f"rows_filled={report.rows_filled} unmatched={report.rows_unmatched} "
          f"cells={report.cells_written}")
    if report.rows_filled < 15:
        _fail(f"expected >=15 GMALG1 rows filled, got {report.rows_filled}")
    if report.cells_written.get("Antenna azimuth_TSSR", 0) < 15:
        _fail("azimuth not filled for enough rows")

    # 4. Read back: GMALG1 rows must have TSSR azimuth filled and Status blank.
    mwb = load_workbook(io.BytesIO(out))
    ws = mwb["RF Parameter"]
    hdr = {c.value: c.column for c in ws[1]}
    checked = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, hdr["Site Name"]).value != "GMALG1":
            continue
        az = ws.cell(r, hdr["Antenna azimuth_TSSR"]).value
        status = ws.cell(r, hdr["TSSR Status"]).value
        if az in (None, ""):
            _fail(f"row {r}: azimuth_TSSR not filled")
        if status not in (None, ""):
            _fail(f"row {r}: TSSR Status should be blank, got {status!r}")
        checked += 1
    print(f"Verified {checked} GMALG1 rows (azimuth filled, status blank)")
    if checked < 15:
        _fail(f"expected to verify >=15 rows, got {checked}")

    # 5. Unit checks for the matching/reconciliation helpers.
    from pdf_excel_merger.tssr_fill import antenna_num, suffix_ordinal
    if antenna_num("Antenna 2") != 2 or antenna_num("8-Port_Antenna") is not None:
        _fail("antenna_num parsing wrong")
    cases = {"1800|L1800_P": 1, "L850_3": 3, "UMTS_Y": 3, "L2.6_sec2": 2,
             "900_C": 3, "L850_1": 1}
    for tok, exp in cases.items():
        if suffix_ordinal(tok) != exp:
            _fail(f"suffix_ordinal({tok!r})={suffix_ordinal(tok)}, expected {exp}")
    # GMALG1 is a clean site -> no sector conflicts expected.
    if report.conflicts:
        _fail(f"GMALG1 should have no conflicts, got {report.conflicts}")
    print("Helper checks passed (antenna_num, suffix_ordinal, 0 conflicts)")

    print("\nALL TSSR CHECKS PASSED ✅")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
