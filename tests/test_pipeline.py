"""End-to-end test of the network-sites workflow (no Streamlit needed).

Generates samples, loads mixed PDF/Excel single-site documents, auto-detects the
site-key columns, matches each document to a master-sheet row, fills the mapped
fields, and verifies the written cells. Run:  python -m tests.test_pipeline
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import make_samples  # noqa: E402
from pdf_excel_merger import excel_reader, mapping as mapping_mod, merger  # noqa: E402
from pdf_excel_merger import site_matching  # noqa: E402
from pdf_excel_merger.pdf_extractor import SOURCE_COLUMN  # noqa: E402
from pdf_excel_merger.source_loader import load_sources  # noqa: E402

SAMPLES = ROOT / "sample_data"


def _fail(msg: str) -> None:
    raise AssertionError(msg)


def main() -> int:
    make_samples.main()

    # 1. Load mixed single-site documents (PDF key/value, PDF records, Excel).
    docs = [
        "site_COL-001.pdf", "site_GAL-002.pdf", "site_KAN-003.pdf",
        "site_MAT-004.xlsx", "site_XYZ-999.pdf",
    ]
    files = [(name, (SAMPLES / name).read_bytes()) for name in docs]
    loaded = load_sources(files, orientation="auto", pdf_strategy="auto")
    df = loaded.combined
    print(f"Loaded {len(df)} site records, columns: {list(df.columns)}")
    for w in loaded.warnings:
        print("  warn:", w)
    if len(df) != 5:
        _fail(f"expected 5 site records, got {len(df)}")
    for col in ["Site ID", "Power (kW)", "Battery Backup (h)", "Status"]:
        if col not in df.columns:
            _fail(f"missing source column: {col}")

    # 2. Read master + detect key columns.
    master_bytes = (SAMPLES / "sites_master.xlsx").read_bytes()
    info = excel_reader.read_template(io.BytesIO(master_bytes), "Sites", 1)
    print(f"Master columns: {info.column_names}")

    main_key = site_matching.detect_key_column(info.column_names)
    source_key = site_matching.detect_key_column(
        [c for c in df.columns if c != SOURCE_COLUMN]
    )
    print(f"Detected key columns -> master: {main_key!r}, source: {source_key!r}")
    if main_key != "Site ID" or source_key != "Site ID":
        _fail("key-column detection failed")

    # 3. Auto-map the data fields (key column handled separately).
    source_cols = [c for c in df.columns if c not in (SOURCE_COLUMN, source_key)]
    targets = [c for c in info.column_names if c != main_key]
    suggestions = mapping_mod.suggest_mapping(targets, source_cols)
    field_map = {t: s.source for t, s in suggestions.items()}
    print("Field mapping:")
    for t, s in suggestions.items():
        print(f"  {t:20s} <- {s.source}")
    for tgt in ["Power (kW)", "Battery Backup (h)", "Status"]:
        if field_map.get(tgt) != tgt:
            _fail(f"field '{tgt}' mapped to {field_map.get(tgt)!r}, expected itself")

    # 4. Preview matches.
    preview = merger.preview_key_matches(
        master_bytes, info, df, main_key_col=main_key, source_key_col=source_key
    )
    print("\nMatch preview:")
    print(preview.to_string(index=False))

    # 5. Merge by key (skip unmatched sites).
    out_bytes, report = merger.merge_by_key(
        master_bytes, info, df, field_map,
        main_key_col=main_key, source_key_col=source_key, on_unmatched="skip",
    )
    print(f"\nUpdated {report.rows_updated} rows, "
          f"unmatched {len(report.unmatched)}, appended {report.rows_appended}")
    if report.rows_updated != 4:
        _fail(f"expected 4 rows updated, got {report.rows_updated}")
    if len(report.unmatched) != 1 or report.unmatched[0]["site_key"] != "XYZ-999":
        _fail(f"expected XYZ-999 unmatched, got {report.unmatched}")

    # 6. Verify the master sheet was filled in the right rows.
    wb = load_workbook(io.BytesIO(out_bytes))
    ws = wb["Sites"]
    by_id = {}
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid:
            by_id[sid] = [ws.cell(r, c).value for c in range(1, 7)]

    # COL-001: Power 5.5, Battery 8, Status Active
    col = by_id["COL-001"]
    print(f"\nCOL-001 row -> {col}")
    if abs(float(col[3]) - 5.5) > 1e-6:
        _fail(f"COL-001 Power should be 5.5, got {col[3]!r}")
    if col[4] != 8:
        _fail(f"COL-001 Battery should be 8, got {col[4]!r}")
    if col[5] != "Active":
        _fail(f"COL-001 Status should be 'Active', got {col[5]!r}")

    # KAN-003 came from a records-style PDF.
    kan = by_id["KAN-003"]
    print(f"KAN-003 row -> {kan}")
    if abs(float(kan[3]) - 4.2) > 1e-6 or kan[5] != "Maintenance":
        _fail(f"KAN-003 not filled correctly: {kan}")

    # MAT-004 came from an already-converted Excel.
    mat = by_id["MAT-004"]
    print(f"MAT-004 row -> {mat}")
    if abs(float(mat[3]) - 3.8) > 1e-6 or mat[4] != 4:
        _fail(f"MAT-004 not filled correctly: {mat}")

    # JAF-005 had no document -> must remain blank.
    jaf = by_id["JAF-005"]
    print(f"JAF-005 row -> {jaf}")
    if any(jaf[i] not in (None, "") for i in (3, 4, 5)):
        _fail(f"JAF-005 should stay blank, got {jaf}")

    print("\nALL CHECKS PASSED ✅")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
